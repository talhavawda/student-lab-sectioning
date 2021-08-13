"""
    Python Notes on working with Excel files and using DataFrames

    Sources:
        https://www.datacamp.com/community/tutorials/python-excel-tutorial
"""

"""
    Write Pandas DataFrames to Excel Files
"""

# Install XlsxWriter if want to write to multiple worksheets

#Specify/Define a Writer (in which the data frame output will be saved using an ExcelWriter object to output the DataFrame.)
writer = pandas.ExcelWriter('example.xlsx', engine='xlsxwriter')

# Write your DataFrame to a file
# yourData is a dataframe that you are interested in writing as an excel file
# pass in the writer variable to the to_excel() function, and you also specify the sheet name. This way, you add a sheet with the data to an existing workbook, which could have many worksheets in a workbook: you can use the ExcelWriter to save multiple, different DataFrames to one workbook having multiple sheets.
yourData.to_excel(writer, 'Sheet1')

# Save the result
writer.save()


"""
    Read and Write Excel Files with Openpyxl
"""

# Install openpyxl

from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('example.xlsx')

# Get sheet names
wbSheetNames = wb.sheetnames

# Get a sheet by name
sheet = wb['Sheet1']

# Get a sheet by name
sheet = wb['Sheet1']

# Print the sheet title
print('Sheet Title:',sheet.title)

# Get currently active sheet
anotherSheet = wb.active


# Retrieve the value of a certain cell
print(sheet['A1'].value)

# Select element 'B3' of your sheet
c = sheet['B3']

# Retrieve the row number of your element
print('Row No.:', c.row)

# Retrieve the column number of your element - A will be 1, B will be 2, etc.
print('Column Letter:', c.column)

# Retrieve the coordinates of the cell
print('Coordinates of cell:', c.coordinate)

# Retrieve cell value
print(sheet.cell(row=1, column=2).value)


# Print out values in column 2 using a loop
for i in range(1, 4):
    print(i, sheet.cell(row=i, column=2).value)



# PRINT OUT ENTIRE FILE (ALL ROWS)
# Print row per row
for cellObj in sheet['A1':'C3']:
    for cell in cellObj:
        #print(cell.coordinate, cell.value)
        print(cell.coordinate, ": ", cell.value, sep="", end="\t")
    #print('--- END ---')
    print()


# Retrieve the maximum amount of rows
print('Max Rows:', sheet.max_row)

# Retrieve the maximum amount of columns
print('Max Columns:', sheet.max_column)


# Import relevant modules from `openpyxl.utils`
from openpyxl.utils import get_column_letter, column_index_from_string

# Return 'A'
print('Column Letter:', get_column_letter(1))

# Return '1'
print('Column Index:', column_index_from_string('A')



"""
    Using Pandas' DataFrame
"""

# Import `pandas`
import pandas

# Convert Sheet to DataFrame
df = pandas.DataFrame(sheet.values)

# APPEND OR WRITE VALUES BACK TO AN EXCEL FILE
# Import `dataframe_to_rows`
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import *

# Initialize a workbook
wb = Workbook()

# Get the worksheet in the active workbook
ws = wb.active

# Append the rows of the DataFrame to your worksheet
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

#First and last few rows
print(coursesDF.head)
print(coursesDF.tail)


"""
    Using pyexcel
"""

"""
pyexcel is a Python Wrapper that provides a single API interface for reading, manipulating, 
and writing data in .csv, .ods, .xls, .xlsx, and .xlsm files. With pyexcel, the data in excel 
files can be turned into an array or dict format with minimal code.
"""

# Import `pyexcel`
import pyexcel

""" Reading Files with pyexcel"""

# Convert your excel data into an array format (Get an array from the data)
my_array = pyexcel.get_array(file_name="test.xls")


# Convert your excel data into an ordered dictionary of lists:
# Import `OrderedDict` module
from pyexcel._compact import OrderedDict
# Get your data in an ordered dictionary of lists
my_dict = pyexcel.get_dict(file_name="test.xls", name_columns_by_row=0)

# Get a dictionary of two-dimensional arrays:
# Get your data in a dictionary of 2D arrays
book_dict = pyexcel.get_book_dict(file_name="test.xls")

# the above two outputs, my_dict, and book_dict, can be converted to a DataFrame using pd.DataFrame()


""" Writing Files with pyexcel"""
# Save the array to a file
pyexcel.save_as(array=data, dest_file_name="array_data.xls")

# Save 2d array dictionary to a file
pyexcel.save_book_as(bookdict=2d_array_dictionary, dest_file_name="2d_array_data.xls")

